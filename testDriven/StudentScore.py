import openpyxl

path = "C:\\Users\\muhds\\Desktop\\test_score.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]
rows = sheet.max_row


scores = []
#from rows to end of rows
for score in range(2, rows+1):
    StudentScore = sheet.cell(score, 1).value
    scores.append(int(StudentScore))
print(max(scores))
print(sum(scores))
print(scores)

markah = 0
for result in scores:
    markah = markah + int(result)
print(markah)
