import time

from PageObjects.FimmSignUp import SignUp
from PageObjects.LoginPage import LoginPage
import openpyxl
def login(self, username, password, role_index):
    loginPage = LoginPage(self.driver)
    loginPage.InsName(username)
    loginPage.InsPassword(password)
    loginPage.clickSubmit()
    self.wait()
    self.SelectIndex(loginPage.SelectRole(), role_index)
    loginPage.clickSubmit()
    time.sleep(5)


def data_excel_Nric():
    path = "C:\\Users\\muhds\\Desktop\\RegisterConsultant.xlsx"
    test_list = []
    workbook = openpyxl.load_workbook(path)

    sheet = workbook["NRIC"]

    rows = sheet.max_row

    for r in range(2, rows+1):
        NRIC = str(sheet.cell(r,1).value)
        email = str(sheet.cell(r,2).value)

        test_list.append((NRIC, email))
    return test_list


def data_excel_Passport():
    path = "C:\\Users\\muhds\\Desktop\\RegisterConsultant.xlsx"
    test_list = []
    workbook = openpyxl.load_workbook(path)

    sheet = workbook["Passport"]

    rows = sheet.max_row

    for r in range(2, rows+1):
        Passport = sheet.cell(r,1).value
        ExpiryDate = sheet.cell(r,2).value
        Email = sheet.cell(r, 3).value

        test_list.append((Passport, ExpiryDate, Email))

    return test_list




