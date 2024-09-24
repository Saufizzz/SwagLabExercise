import inspect
import time
import pytest
import logging
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
import openpyxl

@pytest.mark.usefixtures("setup_and_teardown")
class BaseClass:

    def ScrollPage(self, pixels=500, delay= 10):
        time.sleep(delay)
        self.driver.execute_script(f"window.scrollBy(0,{pixels});")

    def WaitElementPresent(self, locator):
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(locator))

    def WaitElementClickable(self,locator):
        WebDriverWait(self.driver,10).until(EC.element_to_be_clickable(locator))

    def ActionMove(self,locator1):
        return ActionChains(self.driver).move_to_element(locator1)

    def SelectIndex(self,locator,index):
        return Select(locator).select_by_index(index)

    def SelectByText(self,locator, text):
        return Select(locator).select_by_visible_text(text)

    def SelectByValue(self,locator,value):
        return Select(locator).select_by_value(value)


    def wait(self):
        return self.driver.implicitly_wait(10)

    def getLogger(self):
        # This line retrieves the name of the calling function.
        loggerName = inspect.stack()[1][3]

        # This line creates or retrieves a logger object using the name of the calling function.
        logger = logging.getLogger(loggerName)

        # Check if the logger has handlers already to avoid duplicate logs.
        if not logger.handlers:
            # Create a FileHandler object to write log records to a file.
            filehandler = logging.FileHandler("logfile.log")

            # Create a Formatter object with the specified format.
            formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")

            # Set the formatter to the FileHandler.
            filehandler.setFormatter(formatter)

            # Add the FileHandler to the logger.
            logger.addHandler(filehandler)

            # Optionally, set the logging level. For example, to DEBUG.
            logger.setLevel(logging.DEBUG)

        return logger



    # def getLogger(self):
    #     #This line retrieves the name of the calling function (the function that called the function containing this code). It uses the inspect module to inspect the call stack and retrieves the name of the function at index 1, which is the immediate caller, and then selects the third element of the tuple returned by inspect.stack(), which is the name of the function.
    #     loggerName = inspect.stack()[1][3]
    #     #This line creates a logger object using the name of the calling function obtained in the previous line. It uses the getLogger function from the logging module to get or create a logger with the specified name.
    #     logger = logging.getLogger(loggerName)
    #     #This line creates a FileHandler object named filehandler, which will be responsible for writing log records to a file named "logfile.log".
    #     filehandler = logging.FileHandler("logfile.log")
    #     #This line creates a Formatter object named formatter with a specified format. The format string "%(asctime)s : %(levelname)s : %(name)s : %(message)s" specifies the format for log records. It includes placeholders like %(asctime)s for the time of the log message, %(levelname)s for the log level, %(name)s for the logger name, and %(message)s for the log message itself.
    #     formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
    #     #This line sets the formatter (formatter) created in the previous line to the FileHandler object (filehandler). This ensures that log records written by this handler will be formatted according to the specified format
    #     filehandler.setFormatter(formatter)



    def ScrollAndLoadAllProducts(self):
        SCROLL_PAUSE_TIME = 1  # Adjust this pause time as necessary
        last_height = self.driver.execute_script("return document.body.scrollHeight")

        while True:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(SCROLL_PAUSE_TIME)

            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

    def retry_action(action, max_attempts=3, delay=1):
        attempts = 0
        while attempts < max_attempts:
            try:
                # Attempt the action
                result = action()
                return result  # Return the result if successful
            except Exception as e:
                # Handle the exception (or you can just pass)
                print(f"Attempt {attempts + 1} failed:", e)
                # Increment attempts counter
                attempts += 1
                # Wait for a short delay before retrying
                time.sleep(delay)
        # If max_attempts reached without success, raise an exception
        raise Exception("Max attempts reached without success")



    def ReadExcelData(self,path,sheet_name):
        final_list = []
        book = openpyxl.load_workbook(path)
        sheet = book[sheet_name]
        total_rows = sheet.max_row
        total_column = sheet.max_column
        for r in range(2,total_rows+1):
            row_list = []
            for c in range(1,total_column+1):
                row_list.append(sheet.cell(row=r, column=c).value)
            final_list.append(row_list)

        return final_list


