import openpyxl
import pytest

def test_data_excel_Nric():
    path = "C:\\Users\\muhds\\Desktop\\RegisterConsultant.xlsx"
    test_list = []
    workbook = openpyxl.load_workbook(path)

    sheet = workbook["NRIC"]

    rows = sheet.max_row

    for r in range(2, rows+1):
        NRIC = sheet.cell(r,1).value
        email = sheet.cell(r,2).value

        test_list.append((NRIC, email))
        print(test_list)
        # return test_list