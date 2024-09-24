import openpyxl

# to open excel workbook with path, need to add double backslash
book = openpyxl.load_workbook("C:\\Users\\muhds\\Desktop\\FIMMTemplate.xlsx")
# This line gets the active sheet from the opened workbook (book). In Excel, a workbook can contain multiple sheets.
# This line retrieves the currently active sheet.
sheet = book.active
# This line assigns the cell variable to the cell method of the sheet object. This method is used to access
# individual cells within the sheet.
cell = sheet.cell
# This line prints the value of the cell at row A and column 1 in the active sheet (sheet). It accesses the cell
# using indexing notation (['A1']) and then retrieves its value using the .value attribute. This will print the
# content of the cell located at cell A1 in the active sheet.
print(sheet['A1'].value)

'''for i in range(2,sheet.max_column+1): #to get column
    for j in range(1,sheet.max_row+1): #to get row
        dict[(sheet.cell(row=j, column= 2).value)] = sheet.cell(row=j, column=i).value

print(dict)'''

my_dict = {}  # Create an empty dictionary to store values

# Loop over columns
for i in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=i).value  # Get the header from the first row
    values = []  # Create an empty list to store values for this header

    # Loop over rows starting from the second row
    for j in range(2, sheet.max_row + 1):
        value = sheet.cell(row=j, column=i).value  # Get the value from the current cell
        values.append(value)  # Append the value to the list of values for this header

    my_dict[header] = values  # Associate the header with its list of values in the dictionary

# Print the dictionary
for header, values in my_dict.items():
    print(f"{header}: {values}")

