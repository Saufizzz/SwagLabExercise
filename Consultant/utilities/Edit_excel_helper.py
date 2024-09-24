import openpyxl
from openpyxl.reader.excel import load_workbook


def load_excel_file(file_path):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return workbook, sheet


def modify_excel_file(sheet, modifications):
    # Apply modifications to the sheet
    for (row, col, new_value) in modifications:
        sheet.cell(row=row, column=col).value = new_value

