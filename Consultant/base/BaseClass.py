import inspect
import os #Imports the os module, which provides functions for interacting with the operating system
import tempfile #Imports the tempfile module, which is used to create temporary files.
import time
import openpyxl
import pytest
import logging

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook #Imports the Workbook class from the openpyxl library, which is used to create and manipulate Excel files.
from selenium.common import NoSuchElementException, TimeoutException, StaleElementReferenceException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select

@pytest.mark.usefixtures("setup_and_teardown")
class BaseClass:

    def ScrollPage(self, pixels, delay= 10):
        time.sleep(delay)
        self.driver.execute_script(f"window.scrollBy(0,{pixels});")

    def WaitElementPresent(self,locator):
        WebDriverWait(self.driver,10).until(EC.presence_of_element_located(locator))

    def WaitElementClickable(self,locator):
        WebDriverWait(self.driver,10).until(EC.element_to_be_clickable(locator))

    def ActionMove(self,locator1):
        return ActionChains(self.driver).move_to_element(locator1)

    def select_option(self, locator, option, select_by='index'):
        """
        Select an option from a dropdown.

        :param locator: Tuple containing the By strategy and the locator.
        :param option: The option to select. It can be an index (int), visible text (str), or value (str).
        :param select_by: Criteria to select option ('index', 'text', or 'value').
        """
        try:
            # Locate the dropdown element
            element = self.driver.find_element(*locator)
            select = Select(element)

            if select_by == 'index':
                # Select by index
                if isinstance(option, int):
                    select.select_by_index(option)
                else:
                    raise ValueError("Option must be an integer for index selection.")
            elif select_by == 'text':
                # Select by visible text
                if isinstance(option, str):
                    select.select_by_visible_text(option)
                else:
                    raise ValueError("Option must be a string for text selection.")
            elif select_by == 'value':
                # Select by value attribute
                if isinstance(option, str):
                    select.select_by_value(option)
                else:
                    raise ValueError("Option must be a string for value selection.")
            else:
                raise ValueError("Invalid value for select_by. Use 'index', 'text', or 'value'.")

        except NoSuchElementException:
            print(f"Element with locator {locator} not found.")
        except TimeoutException:
            print(f"Element with locator {locator} was not interactable.")
        except Exception as e:
            print(f"An error occurred: {e}")



    def wait(self):
        return self.driver.implicitly_wait(10)

    def getLogger(self):
        # This line retrieves the name of the calling function.
        loggerName = inspect.stack()[1][3]

        # This line creates or retrieves a logger object using the name of the calling function.
        logger = logging.getLogger(loggerName)

        # Check if the logger has handlers already to avoid duplicate logs.
        if not logger.handlers:
            # Create a FileHandler object to write log records to a file.
            filehandler = logging.FileHandler("logfile.log")

            # Create a Formatter object with the specified format.
            formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")

            # Set the formatter to the FileHandler.
            filehandler.setFormatter(formatter)

            # Add the FileHandler to the logger.
            logger.addHandler(filehandler)

            # Optionally, set the logging level. For example, to DEBUG.
            logger.setLevel(logging.DEBUG)

        return logger



    # def getLogger(self):
    #     #This line retrieves the name of the calling function (the function that called the function containing this code). It uses the inspect module to inspect the call stack and retrieves the name of the function at index 1, which is the immediate caller, and then selects the third element of the tuple returned by inspect.stack(), which is the name of the function.
    #     loggerName = inspect.stack()[1][3]
    #     #This line creates a logger object using the name of the calling function obtained in the previous line. It uses the getLogger function from the logging module to get or create a logger with the specified name.
    #     logger = logging.getLogger(loggerName)
    #     #This line creates a FileHandler object named filehandler, which will be responsible for writing log records to a file named "logfile.log".
    #     filehandler = logging.FileHandler("logfile.log")
    #     #This line creates a Formatter object named formatter with a specified format. The format string "%(asctime)s : %(levelname)s : %(name)s : %(message)s" specifies the format for log records. It includes placeholders like %(asctime)s for the time of the log message, %(levelname)s for the log level, %(name)s for the logger name, and %(message)s for the log message itself.
    #     formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
    #     #This line sets the formatter (formatter) created in the previous line to the FileHandler object (filehandler). This ensures that log records written by this handler will be formatted according to the specified format
    #     filehandler.setFormatter(formatter)



    def ScrollAndLoadAllProducts(self):
        SCROLL_PAUSE_TIME = 3  # Adjust this pause time as necessary
        last_height = self.driver.execute_script("return document.body.scrollHeight")

        while True:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(SCROLL_PAUSE_TIME)

            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

    def retry_action(action, max_attempts=3, delay=1):
        attempts = 0
        while attempts < max_attempts:
            try:
                # Attempt the action
                result = action()
                return result  # Return the result if successful
            except Exception as e:
                # Handle the exception (or you can just pass)
                print(f"Attempt {attempts + 1} failed:", e)
                # Increment attempts counter
                attempts += 1
                # Wait for a short delay before retrying
                time.sleep(delay)
        # If max_attempts reached without success, raise an exception
        raise Exception("Max attempts reached without success")



    def read_excel(file_path, sheet_name):
        # Load the workbook (Excel file) from the specified file path
        workbook = openpyxl.load_workbook(file_path)
        # Select the sheet within the workbook based on the provided sheet name
        sheet = workbook[sheet_name]

        # Read the header row (first row) of the Excel sheet
        headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        # Initialize an empty list to store the data rows
        data = []

        # Loop through each row in the sheet starting from the second row to the last row
        for row in range(2, sheet.max_row + 1):
            # Create a tuple containing the values of each cell in the current row
            values = tuple(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1))
            # Append the tuple of values to the data list
            data.append(values)

        # Return the headers and the data
        return headers, data

    def retry_find_element(self, locator, timeout=10, poll_frequency=0.5):
        """
        Retry finding an element in case of StaleElementReferenceException.
        """
        end_time = time.time() + timeout
        while True:
            try:
                element = WebDriverWait(self.driver, poll_frequency).until(
                    EC.presence_of_element_located(locator)
                )
                return element
            except StaleElementReferenceException:
                if time.time() > end_time:
                    raise
                time.sleep(poll_frequency)



