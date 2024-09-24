import inspect
import os #Imports the os module, which provides functions for interacting with the operating system
import tempfile #Imports the tempfile module, which is used to create temporary files.
import time
import openpyxl
import pytest
import logging
from openpyxl.workbook import Workbook #Imports the Workbook class from the openpyxl library, which is used to create and manipulate Excel files.
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select

@pytest.mark.usefixtures("setup_and_teardown")
class BaseClass:

    def ScrollPage(self, pixels, delay= 10):
        time.sleep(delay)
        self.driver.execute_script(f"window.scrollBy(0,{pixels});")

    def WaitElementPresent(self,locator):
        WebDriverWait(self.driver,10).until(EC.presence_of_element_located(locator))

    def WaitElementClickable(self,locator):
        WebDriverWait(self.driver,10).until(EC.element_to_be_clickable(locator))

    def ActionMove(self,locator1):
        return ActionChains(self.driver).move_to_element(locator1)

    def selectIndex(self,locator,index):
        return Select(locator).select_by_index(index)

    def selectByText(self,locator, text):
        return Select(locator).select_by_visible_text(text)


    def wait(self):
        return self.driver.implicitly_wait(10)

    def getLogger(self):
        # This line retrieves the name of the calling function.
        loggerName = inspect.stack()[1][3]

        # This line creates or retrieves a logger object using the name of the calling function.
        logger = logging.getLogger(loggerName)

        # Check if the logger has handlers already to avoid duplicate logs.
        if not logger.handlers:
            # Create a FileHandler object to write log records to a file.
            filehandler = logging.FileHandler("logfile.log")

            # Create a Formatter object with the specified format.
            formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")

            # Set the formatter to the FileHandler.
            filehandler.setFormatter(formatter)

            # Add the FileHandler to the logger.
            logger.addHandler(filehandler)

            # Optionally, set the logging level. For example, to DEBUG.
            logger.setLevel(logging.DEBUG)

        return logger



    # def getLogger(self):
    #     #This line retrieves the name of the calling function (the function that called the function containing this code). It uses the inspect module to inspect the call stack and retrieves the name of the function at index 1, which is the immediate caller, and then selects the third element of the tuple returned by inspect.stack(), which is the name of the function.
    #     loggerName = inspect.stack()[1][3]
    #     #This line creates a logger object using the name of the calling function obtained in the previous line. It uses the getLogger function from the logging module to get or create a logger with the specified name.
    #     logger = logging.getLogger(loggerName)
    #     #This line creates a FileHandler object named filehandler, which will be responsible for writing log records to a file named "logfile.log".
    #     filehandler = logging.FileHandler("logfile.log")
    #     #This line creates a Formatter object named formatter with a specified format. The format string "%(asctime)s : %(levelname)s : %(name)s : %(message)s" specifies the format for log records. It includes placeholders like %(asctime)s for the time of the log message, %(levelname)s for the log level, %(name)s for the logger name, and %(message)s for the log message itself.
    #     formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
    #     #This line sets the formatter (formatter) created in the previous line to the FileHandler object (filehandler). This ensures that log records written by this handler will be formatted according to the specified format
    #     filehandler.setFormatter(formatter)



    def ScrollAndLoadAllProducts(self):
        SCROLL_PAUSE_TIME = 3  # Adjust this pause time as necessary
        last_height = self.driver.execute_script("return document.body.scrollHeight")

        while True:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(SCROLL_PAUSE_TIME)

            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

    def retry_action(action, max_attempts=3, delay=1):
        attempts = 0
        while attempts < max_attempts:
            try:
                # Attempt the action
                result = action()
                return result  # Return the result if successful
            except Exception as e:
                # Handle the exception (or you can just pass)
                print(f"Attempt {attempts + 1} failed:", e)
                # Increment attempts counter
                attempts += 1
                # Wait for a short delay before retrying
                time.sleep(delay)
        # If max_attempts reached without success, raise an exception
        raise Exception("Max attempts reached without success")



    def read_excel(file_path, sheet_name):
        # Load the workbook (Excel file) from the specified file path
        workbook = openpyxl.load_workbook(file_path)
        # Select the sheet within the workbook based on the provided sheet name
        sheet = workbook[sheet_name]

        # Read the header row (first row) of the Excel sheet
        headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        # Initialize an empty list to store the data rows
        data = []

        # Loop through each row in the sheet starting from the second row to the last row
        for row in range(2, sheet.max_row + 1):
            # Create a tuple containing the values of each cell in the current row
            values = tuple(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1))
            # Append the tuple of values to the data list
            data.append(values)

        # Return the headers and the data
        return headers, data

    # Fixture to create an in-memory Excel file
    @pytest.fixture
    def excel_file(self):
        wb = Workbook() # Create a new workbook object
        ws = wb.active # Get the active worksheet (default sheet)
        ws.title = "SampleData" # Set the title of the worksheet
        ws.append(["Column1", "Column2", "Column3"])  # Add a header row with column names
        ws.append(["Data1", "Data2", "Data3"])  # Add a data row
        # You can add more rows as needed using ws.append([...])
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") # Create a temporary file with .xlsx extension
        wb.save(temp_file.name) # Save the workbook to the temporary file
        yield temp_file.name # Yield the temporary file's name for use in tests
        os.remove(temp_file.name) # Remove the temporary file after the test
        # when want to run the test, include excel_file as argument. e.g (self,excel_file)

