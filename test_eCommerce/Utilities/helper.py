import openpyxl
from pageObjectModel.loginPage import LoginPage

# How to use excel data to parametrize python
def excelData():
    path = "C:\\Users\\muhds\\Desktop\\RegisterConsultant.xlsx"  #Dummy Data
    test_list = []
    workbook = openpyxl.load_workbook(path)

    sheet = workbook["Passport"]  #Dummy Data

    rows = sheet.max_row

    for r in range(2, rows+1):
        Passport = sheet.cell(r,1).value  #Dummy Data
        ExpiryDate = sheet.cell(r,2).value  #Dummy Data
        Email = sheet.cell(r, 3).value  #Dummy Data

        test_list.append((Passport, ExpiryDate, Email))  #Dummy Data

    return test_list

def login_user(self, username, password):
    login = LoginPage(self.driver)
    login.insertUsername(username)
    login.insertPassword(password)
    login.clickLoginBtn()
